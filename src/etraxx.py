"""
FSG CCBOM Automation Tool ‚ÄĒ e-traxx variant (logic module)
==========================================================

Variant of bom_automation.py adapted to the e-traxx master BOM schema
(BOMs/BOM_Final.xlsx), which differs from the EFRxx template:

  * Header row is row 2 (row 1 is a category banner).
  * "System" contains full names ("Chassis and Body"), not codes.
  * "Make/Buy" uses the words "Make" / "Buy".
  * Part column is "Part Name"; comments column is "Comment".
  * A boolean column "if Make CCBOM Eintrag erstellt?" marks already-
    uploaded parts ‚ÄĒ treated equivalently to a green row.
  * An "Eingebaut?" boolean gates whether a part is actually on the car;
    set ETRAXX_REQUIRE_INSTALLED=true in .env to only upload installed parts.

All environment configuration lives in `src/env.py`; this module contains
the constants, helpers, and main automation loop.
"""

import csv
import glob
import os
import sys
import time
from datetime import datetime

import openpyxl
import pandas as pd
from playwright.sync_api import sync_playwright

from src.env import (
    BOM_URL,
    BOMS_DIR,
    DEFAULT_FILE,
    DEFAULT_SYSTEM,
    DRY_RUN,
    DRY_RUN_HOLD_MS,
    FSG_PASSWORD,
    FSG_USERNAME,
    LOG_FILE,
    LOGIN_URL,
    REQUIRE_INSTALLED,
    TEAM_ID,
    TEST_LIMIT,
    TEST_MODE,
)

# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
# Constants
# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ

SKIP_COLORS = {
    "FF00FF00",
    "0000FF00",  # Green ‚ÄĒ already uploaded
    "FFFF0000",
    "00FF0000",  # Red   ‚ÄĒ do not upload
}

SYSTEM_MAP = {
    "AT": "AT - Autonomous System",
    "BR": "BR - Brake System",
    "DT": "DT - Drivetrain",
    "ET": "ET - Engine and Tractive System",
    "FR": "FR - Chassis and Body",
    "LV": "LV - Grounded Low Voltage System",
    "MS": "MS - Miscellaneous Fit and Finish",
    "ST": "ST - Steering System",
    "SU": "SU - Suspension System",
    "WT": "WT - Wheels, Wheel Bearings and Tires",
}

# Full name (as written in BOM_Final.xlsx) ‚Üí system code
SYSTEM_NAME_TO_CODE = {
    "autonomous system": "AT",
    "brake system": "BR",
    "drivetrain": "DT",
    "engine and tractive system": "ET",
    "chassis and body": "FR",
    "grounded low voltage system": "LV",
    "miscellaneous fit and finish": "MS",
    "steering system": "ST",
    "suspension system": "SU",
    "wheels, wheel bearings and tires": "WT",
    "wheels wheel bearings and tires": "WT",
}

ASSEMBLY_REMAP = {
    "brake caliper": "Calipers",
    "brake calipers": "Calipers",
    "caliper": "Calipers",
    "reservoire": "Brake Master Cylinder",
    "reservoir": "Brake Master Cylinder",
    "resovoir": "Brake Master Cylinder",
    "fitting screw": "Fasteners",
    "fastener": "Fasteners",
    "screws": "Fasteners",
    "bolts": "Fasteners",
    "brake disc": "Brake Discs",
    "brake disk": "Brake Discs",
    "brake pad": "Brake Pads",
    "brake line": "Brake Lines",
    "master cylinder": "Brake Master Cylinder",
    "damper": "Dampers",
    "spring": "Springs",
    "pushrod": "Pushrods",
    "rocker": "Rockers",
    "a-arm": "A-Arms",
    "chain": "Chain",
    "sprocket": "Sprockets",
    "differential": "Differential",
    "half shaft": "Half Shafts",
    "halfshaft": "Half Shafts",
    "steering rack": "Steering Rack",
    "tie rod": "Tie Rods",
    "steering wheel": "Steering Wheel",
    "tire": "Tires",
    "tyre": "Tires",
    "wheel bearing": "Wheel Bearings",
    "rim": "Wheels",
    "wheel": "Wheels",
}

# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
# Logging
# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ


def log(message: str, status: str = "INFO") -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [{status}] {message}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


ERROR_CSV = os.getenv("ERROR_CSV", "bom_errors.csv")


def log_error_csv(item: dict, reason: str) -> None:
    path = ERROR_CSV
    new_file = not os.path.isfile(path)
    with open(path, "a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(
                [
                    "timestamp",
                    "row",
                    "system",
                    "assembly",
                    "subassembly",
                    "part",
                    "makebuy",
                    "quantity",
                    "comments",
                    "reason",
                ]
            )
        w.writerow(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                item.get("row", ""),
                item.get("system", ""),
                item.get("assembly", ""),
                item.get("subassembly", ""),
                item.get("part", ""),
                item.get("makebuy", ""),
                item.get("quantity", ""),
                item.get("comments", ""),
                reason,
            ]
        )


# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
# Excel helpers
# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ


def get_cell_color(sheet, row: int, col: int = 1) -> str | None:
    try:
        fill = sheet.cell(row=row, column=col).fill
        if fill.patternType is None:
            return None
        idx = fill.start_color.index
        return str(idx).upper() if idx and idx != "00000000" else None
    except Exception:
        return None


def should_skip_color(sheet, row: int) -> str | None:
    color = get_cell_color(sheet, row)
    if color is None:
        return None
    if color in SKIP_COLORS:
        if "FF00" in color or "00FF00" in color.replace("FF", "", 1):
            return "green (already uploaded)"
        if "FF0000" in color:
            return "red (do not upload)"
        return f"skipped colour ({color})"
    return None


def as_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("true", "1", "yes", "x", "ja")


# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
# Fuzzy dropdown matching
# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ


def snapshot_options(page, selector: str) -> list[str]:
    try:
        return page.eval_on_selector(
            selector,
            "el => Array.from(el.options).map(o => o.text)",
        )
    except Exception:
        return []


def _pick(options: list[str], target: str) -> str | None:
    resolved = ASSEMBLY_REMAP.get(target.lower().strip(), target)
    resolved_lower = resolved.lower().strip()
    if resolved in options:
        return resolved
    for opt in options:
        if opt.lower().strip() == resolved_lower:
            return opt
    for opt in options:
        ol = opt.lower().strip()
        if resolved_lower in ol or ol in resolved_lower:
            return opt
    return None


def wait_for_options(
    page,
    selector: str,
    expected: str | None = None,
    timeout_ms: int = 15000,
    poll_ms: int = 200,
    previous: list[str] | None = None,
) -> list[str]:
    # poll until the target is in the dropdown, period. slow but reliable.
    # if no target given, just wait for any populated list.
    deadline = time.time() + timeout_ms / 1000
    options: list[str] = []
    while time.time() < deadline:
        options = snapshot_options(page, selector)
        real = [o for o in options if o and o.strip()]
        if expected is None:
            if real and real[0].strip().lower() not in ("", "select", "-"):
                return options
        elif real and _pick(options, expected) is not None:
            return options
        page.wait_for_timeout(poll_ms)
    return options


def fuzzy_select(page, selector: str, target: str) -> bool:
    # wait (again) for the target to be present, then select it.
    # retries the select itself in case AJAX wipes the list between
    # snapshot and select_option.
    deadline = time.time() + 15
    while time.time() < deadline:
        options = snapshot_options(page, selector)
        choice = _pick(options, target)
        if choice is not None:
            try:
                page.locator(selector).select_option(label=choice)
                return True
            except Exception:
                page.wait_for_timeout(200)
                continue
        page.wait_for_timeout(200)
    return False


# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
# File selection
# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ


def discover_excel_files() -> list[str]:
    search_dir = os.path.join(os.getcwd(), BOMS_DIR)
    if not os.path.isdir(search_dir):
        os.makedirs(search_dir, exist_ok=True)
        log(f"Created '{BOMS_DIR}/' directory. Place your Excel files there.", "WARN")
        return []
    return sorted(glob.glob(os.path.join(search_dir, "*.xlsx")))


def select_file() -> str:
    # Prefer the configured default file if present.
    default_path = os.path.join(os.getcwd(), BOMS_DIR, DEFAULT_FILE)
    if os.path.isfile(default_path):
        log(f"Using default e-traxx file: {DEFAULT_FILE}")
        return default_path

    files = discover_excel_files()
    if not files:
        print(f"\n  No .xlsx files found in '{BOMS_DIR}/'.\n")
        sys.exit(1)
    print(f"\nExcel files in '{BOMS_DIR}/':")
    for i, f in enumerate(files):
        print(f"  {i + 1}. {os.path.basename(f)}")
    while True:
        try:
            choice = int(input("\nSelect a file (number): ")) - 1
            if 0 <= choice < len(files):
                return files[choice]
        except ValueError:
            pass
        print("Invalid ‚ÄĒ please enter a valid number.")


# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
# Main
# ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ


def main() -> None:
    print("""
    ‚Ėď‚Ėą‚Ėą‚Ėą‚Ėą‚Ėą‚ĖĄ‚ĖĄ‚ĖĄ‚Ėą‚Ėą‚Ėą‚Ėą‚Ėą‚Ėď ‚Ėą‚Ėą‚ĖÄ‚Ėą‚Ėą‚Ėą   ‚ĖĄ‚ĖĄ‚ĖĄ      ‚Ėí‚Ėą‚Ėą   ‚Ėą‚Ėą‚Ėí‚Ėí‚Ėą‚Ėą   ‚Ėą‚Ėą‚Ėí
    ‚Ėď‚Ėą   ‚ĖÄ‚Ėď  ‚Ėą‚Ėą‚Ėí ‚Ėď‚Ėí‚Ėď‚Ėą‚Ėą ‚Ėí ‚Ėą‚Ėą‚Ėí‚Ėí‚Ėą‚Ėą‚Ėą‚Ėą‚ĖĄ    ‚Ėí‚Ėí ‚Ėą ‚Ėą ‚Ėí‚ĖĎ‚Ėí‚Ėí ‚Ėą ‚Ėą ‚Ėí‚ĖĎ
    ‚Ėí‚Ėą‚Ėą‚Ėą  ‚Ėí ‚Ėď‚Ėą‚Ėą‚ĖĎ ‚Ėí‚ĖĎ‚Ėď‚Ėą‚Ėą ‚ĖĎ‚ĖĄ‚Ėą ‚Ėí‚Ėí‚Ėą‚Ėą  ‚ĖÄ‚Ėą‚ĖĄ  ‚ĖĎ‚ĖĎ  ‚Ėą   ‚ĖĎ‚ĖĎ‚ĖĎ  ‚Ėą   ‚ĖĎ
    ‚Ėí‚Ėď‚Ėą  ‚ĖĄ‚ĖĎ ‚Ėď‚Ėą‚Ėą‚Ėď ‚ĖĎ ‚Ėí‚Ėą‚Ėą‚ĖÄ‚ĖÄ‚Ėą‚ĖĄ  ‚ĖĎ‚Ėą‚Ėą‚ĖĄ‚ĖĄ‚ĖĄ‚ĖĄ‚Ėą‚Ėą  ‚ĖĎ ‚Ėą ‚Ėą ‚Ėí  ‚ĖĎ ‚Ėą ‚Ėą ‚Ėí
    ‚ĖĎ‚Ėí‚Ėą‚Ėą‚Ėą‚Ėą‚Ėí ‚Ėí‚Ėą‚Ėą‚Ėí ‚ĖĎ ‚ĖĎ‚Ėą‚Ėą‚Ėď ‚Ėí‚Ėą‚Ėą‚Ėí ‚Ėď‚Ėą   ‚Ėď‚Ėą‚Ėą‚Ėí‚Ėí‚Ėą‚Ėą‚Ėí ‚Ėí‚Ėą‚Ėą‚Ėí‚Ėí‚Ėą‚Ėą‚Ėí ‚Ėí‚Ėą‚Ėą‚Ėí
    ‚ĖĎ‚ĖĎ ‚Ėí‚ĖĎ ‚ĖĎ ‚Ėí ‚ĖĎ‚ĖĎ   ‚ĖĎ ‚Ėí‚Ėď ‚ĖĎ‚Ėí‚Ėď‚ĖĎ ‚Ėí‚Ėí   ‚Ėď‚Ėí‚Ėą‚ĖĎ‚Ėí‚Ėí ‚ĖĎ ‚ĖĎ‚Ėď ‚ĖĎ‚Ėí‚Ėí ‚ĖĎ ‚ĖĎ‚Ėď ‚ĖĎ
     ‚ĖĎ ‚ĖĎ  ‚ĖĎ   ‚ĖĎ      ‚ĖĎ‚Ėí ‚ĖĎ ‚Ėí‚ĖĎ  ‚Ėí   ‚Ėí‚Ėí ‚ĖĎ‚ĖĎ‚ĖĎ   ‚ĖĎ‚Ėí ‚ĖĎ‚ĖĎ‚ĖĎ   ‚ĖĎ‚Ėí ‚ĖĎ
       ‚ĖĎ    ‚ĖĎ        ‚ĖĎ‚ĖĎ   ‚ĖĎ   ‚ĖĎ   ‚Ėí    ‚ĖĎ    ‚ĖĎ   ‚ĖĎ    ‚ĖĎ
       ‚ĖĎ  ‚ĖĎ           ‚ĖĎ           ‚ĖĎ  ‚ĖĎ ‚ĖĎ    ‚ĖĎ   ‚ĖĎ    ‚ĖĎ
                                                        """)
    print("""
    ‚Ėó‚ĖĄ‚ĖĄ‚ĖĖ  ‚Ėó‚ĖĄ‚ĖĖ ‚Ėó‚ĖĖ  ‚Ėó‚ĖĖ    ‚Ėó‚ĖĄ‚ĖĄ‚ĖĄ‚ĖĖ‚Ėó‚ĖĄ‚ĖĖ  ‚Ėó‚ĖĄ‚ĖĖ ‚Ėó‚ĖĖ
    ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚Ėõ‚Ėö‚Ėě‚Ėú‚ĖĆ      ‚Ėą ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚ĖĆ
    ‚Ėź‚Ėõ‚ĖÄ‚Ėö‚ĖĖ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚ĖĆ  ‚Ėź‚ĖĆ      ‚Ėą ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚ĖĆ ‚Ėź‚ĖĆ‚Ėź‚ĖĆ
    ‚Ėź‚Ėô‚ĖĄ‚Ėě‚Ėė‚ĖĚ‚Ėö‚ĖĄ‚Ėě‚Ėė‚Ėź‚ĖĆ  ‚Ėź‚ĖĆ      ‚Ėą ‚ĖĚ‚Ėö‚ĖĄ‚Ėě‚Ėė‚ĖĚ‚Ėö‚ĖĄ‚Ėě‚Ėė‚Ėź‚Ėô‚ĖĄ‚ĖĄ‚ĖĄ‚ĖĖ""")
    if TEST_MODE:
        log(f"TEST MODE: Only the first {TEST_LIMIT} parts will be processed.", "WARN")
    log(
        f"Config: TEAM_ID={TEAM_ID} TEST_MODE={TEST_MODE} DRY_RUN={DRY_RUN} "
        f"BOMS_DIR={BOMS_DIR} REQUIRE_INSTALLED={REQUIRE_INSTALLED}"
    )

    # ‚ĒÄ‚ĒÄ 1. File ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
    filepath = select_file()
    filename = os.path.basename(filepath)
    log(f"Selected file: {filename}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "BOM" in wb.sheetnames:
        sheet = wb["BOM"]
    else:
        sheet = wb.active
        log(f"Sheet 'BOM' not found ‚ÄĒ using active sheet '{sheet.title}'.", "WARN")

    # Header row is row 2 in the e-traxx template (row 1 is a banner).
    df = pd.read_excel(filepath, sheet_name=sheet.title, header=1)
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Map the e-traxx column names to the internal keys we use below.
    # NOTE: website "part" is filled from NXTeilname (the NX/CAD identifier);
    # Excel "Part Name" (human-readable) is pushed into the website comment.
    COLMAP = {
        "system": "system",
        "assembly": "assembly",
        "sub-assembly": "subassembly",
        "part name": "part_label",
        "nxteilname": "part",
        "make/buy": "makebuy",
        "comment": "comments",
        "quantity": "quantity",
        "eingebaut?": "installed",
        "if make ccbom eintrag erstellt?": "uploaded",
    }
    missing = [
        c
        for c in ("system", "assembly", "part name", "nxteilname")
        if c not in df.columns
    ]
    if missing:
        log(
            f"Missing required columns: {missing}. Available: {list(df.columns)}",
            "ERROR",
        )
        sys.exit(1)

    # Header is on row 2 ‚Üí first data row is Excel row 3.
    HEADER_EXCEL_ROW = 2

    def g(row, key, default=""):
        col = next((c for c, k in COLMAP.items() if k == key), None)
        if col is None or col not in row.index:
            return default
        return row[col]

    # ‚ĒÄ‚ĒÄ 2. System selection ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
    unique_codes: list[str] = []
    seen = set()
    for s in df["system"].dropna().unique():
        code = SYSTEM_NAME_TO_CODE.get(str(s).strip().lower())
        if code and code not in seen:
            seen.add(code)
            unique_codes.append(code)

    if DEFAULT_SYSTEM and DEFAULT_SYSTEM in unique_codes:
        run_system = DEFAULT_SYSTEM
        log(
            f"System filter (from .env): {run_system} ‚ÄĒ "
            f"{SYSTEM_MAP.get(run_system, run_system)}"
        )
    else:
        print("\nSystems found in Excel:")
        for s in unique_codes:
            print(f"  ‚ÄĘ {s:4s}  {SYSTEM_MAP.get(s, s)}")
        run_system = (
            input("\nEnter system code to filter (e.g. 'BR') or 'ALL' for everything: ")
            .strip()
            .upper()
        )

    # ‚ĒÄ‚ĒÄ 3. Filter rows ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
    filtered = []
    skipped_green = 0
    skipped_red = 0
    skipped_example = 0
    skipped_empty = 0
    skipped_uploaded = 0
    skipped_not_installed = 0

    for idx, row in df.iterrows():
        excel_row = idx + HEADER_EXCEL_ROW + 1

        sys_name = str(g(row, "system", "")).strip()
        sys_code = SYSTEM_NAME_TO_CODE.get(sys_name.lower(), "")
        assembly = str(g(row, "assembly", "")).strip()
        subassembly = str(g(row, "subassembly", "")).strip()
        if subassembly.lower() in ("nan", ""):
            subassembly = ""
        part = str(g(row, "part", "")).strip()
        part_label = str(g(row, "part_label", "")).strip()
        quantity = g(row, "quantity", "")
        makebuy = str(g(row, "makebuy", "")).strip().lower()
        comments = str(g(row, "comments", "")).strip()
        installed = as_bool(g(row, "installed", False))
        uploaded = as_bool(g(row, "uploaded", False))

        if run_system != "ALL" and sys_code != run_system:
            continue

        if not sys_code or not part or part.lower() in ("nan", ""):
            skipped_empty += 1
            continue

        if part_label.lower() in ("nan", ""):
            part_label = ""

        combined = f"{part.upper()} {part_label.upper()}"
        if "BEISPIEL" in combined or "EXAMPLE" in combined:
            skipped_example += 1
            continue

        if uploaded:
            skipped_uploaded += 1
            log(
                f"Row {excel_row}: Skipped ‚ÄĒ already uploaded (CCBOM Eintrag erstellt)",
                "SKIP",
            )
            continue

        if REQUIRE_INSTALLED and not installed:
            skipped_not_installed += 1
            continue

        skip_reason = should_skip_color(sheet, excel_row)
        if skip_reason:
            if "green" in skip_reason:
                skipped_green += 1
            elif "red" in skip_reason:
                skipped_red += 1
            log(f"Row {excel_row}: Skipped ‚ÄĒ {skip_reason}", "SKIP")
            continue

        mb = makebuy[0] if makebuy and makebuy[0] in ("m", "b") else "m"

        if comments.lower() in ("nan", ""):
            comments = ""

        # Website comment = Excel "Part Name"; append existing comment if any.
        if part_label:
            comments = f"{part_label} ‚ÄĒ {comments}" if comments else part_label

        qty_str = str(quantity).strip()
        if qty_str.lower() in ("nan", ""):
            qty_str = ""
        elif qty_str.endswith(".0"):
            qty_str = qty_str[:-2]

        filtered.append(
            {
                "row": excel_row,
                "system": sys_code,
                "assembly": assembly,
                "subassembly": subassembly,
                "part": part,
                "makebuy": mb,
                "quantity": qty_str,
                "comments": comments,
            }
        )

    log(
        f"Filtering complete: {len(filtered)} parts to upload "
        f"({skipped_green} green / {skipped_red} red / "
        f"{skipped_uploaded} already uploaded / "
        f"{skipped_not_installed} not installed / "
        f"{skipped_example} example / {skipped_empty} empty skipped)"
    )

    if TEST_MODE and len(filtered) > TEST_LIMIT:
        log(f"Test Mode: limiting {len(filtered)} ‚Üí {TEST_LIMIT} parts")
        filtered = filtered[:TEST_LIMIT]

    if not filtered:
        log("Nothing to upload ‚ÄĒ exiting.")
        sys.exit(0)

    if not (FSG_USERNAME and FSG_PASSWORD):
        print("\nWARNING: FSG_USERNAME and/or FSG_PASSWORD not set.")
        print("The script will open a browser and you will need to log in manually.")
        manual_confirm = input(
            "Type 'YES' to continue in manual login mode, or anything else to abort: "
        ).strip()
        if manual_confirm != "YES":
            log(
                "Aborted by user: credentials missing and manual login declined.",
                "ERROR",
            )
            sys.exit(1)

    print(f"\nReady to upload to TEAM_ID={TEAM_ID}. Parts to upload: {len(filtered)}")
    print(f"TEST_MODE={TEST_MODE} DRY_RUN={DRY_RUN}")
    confirm = input(
        "Type 'YES' to proceed with uploading (or anything else to abort): "
    ).strip()
    if confirm != "YES":
        log("Aborted by user before upload.", "WARN")
        sys.exit(0)

    # ‚ĒÄ‚ĒÄ 4. Browser automation ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        if FSG_USERNAME and FSG_PASSWORD:
            log(f"Logging in as '{FSG_USERNAME}'...")
            page.goto(LOGIN_URL)
            page.fill("#tx-felogin-input-username", FSG_USERNAME)
            page.fill("#tx-felogin-input-password", FSG_PASSWORD)
            page.click('input[name="submit"]')
            page.wait_for_load_state("networkidle")

        page.goto(BOM_URL)
        input(
            "\n  ‚ĒĆ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚Ēź\n"
            "  ‚Ēā  Verify you are logged in and on the BOM page.     ‚Ēā\n"
            "  ‚Ēā  Press ENTER to begin uploading.                   ‚Ēā\n"
            "  ‚ĒĒ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚ĒÄ‚Ēė\n"
        )

        log("Fetching existing parts for deduplication...")
        existing: set[str] = set()
        try:
            data = page.evaluate(
                """() => {
                    try { return $('#bom-table').DataTable().data().toArray(); }
                    catch(e) { return []; }
                }"""
            )
            for r in data:
                if isinstance(r, dict):
                    key = (
                        f"{str(r.get('system', '')).strip()}_"
                        f"{str(r.get('assembly', '')).strip()}_"
                        f"{str(r.get('subassembly', '')).strip()}_"
                        f"{str(r.get('part', '')).strip()}"
                    ).lower()
                    existing.add(key)
            log(f"Found {len(existing)} existing parts on the website.")
        except Exception as e:
            log(f"Could not read existing parts: {e}", "WARN")

        success = 0
        failed = 0
        skipped_dup = 0
        start_time = time.time()

        def close_modal() -> None:
            try:
                cancel = page.get_by_text("Cancel", exact=True)
                if cancel.count():
                    cancel.first.click()
                else:
                    page.keyboard.press("Escape")
                page.wait_for_selector(
                    ".DTE_Action_Create", state="hidden", timeout=3000
                )
            except Exception:
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(500)
                except Exception:
                    pass

        def try_create(item: dict) -> None:
            sys_code = item["system"]
            asm_raw = item["assembly"]
            sub_raw = item["subassembly"]
            part_name = item["part"]

            page.get_by_text("New", exact=True).click()
            page.wait_for_selector(".DTE_Action_Create")

            sys_label = SYSTEM_MAP.get(sys_code, sys_code)
            # jiggle: pick a different system first to force a real change
            # event, then switch to the target. prevents stuck assembly list.
            sys_options = snapshot_options(page, "#DTE_Field_system")
            other = next(
                (
                    o
                    for o in sys_options
                    if o
                    and o.strip()
                    and o.strip() != sys_label
                    and o.strip().lower() not in ("", "select", "-")
                ),
                None,
            )
            if other:
                try:
                    page.locator("#DTE_Field_system").select_option(label=other)
                    page.locator("#DTE_Field_system").dispatch_event("change")
                    page.wait_for_timeout(300)
                except Exception:
                    pass
            if not fuzzy_select(page, "#DTE_Field_system", sys_label):
                raise RuntimeError(f"System '{sys_label}' not found in dropdown")
            page.locator("#DTE_Field_system").dispatch_event("change")
            wait_for_options(page, "#DTE_Field_assembly", asm_raw)

            if not fuzzy_select(page, "#DTE_Field_assembly", asm_raw):
                raise RuntimeError(f"Assembly '{asm_raw}' not found in dropdown")
            page.locator("#DTE_Field_assembly").dispatch_event("change")

            # probe whichever id the live DOM exposes ‚ÄĒ the e-traxx template
            # has drifted between "subassembly" and "sub_assembly"
            sub_sel = None
            for candidate in ("#DTE_Field_sub_assembly", "#DTE_Field_subassembly"):
                if page.locator(candidate).count():
                    sub_sel = candidate
                    break

            if sub_raw and sub_sel:
                # only wait for the dropdown to *populate*, not for our specific
                # target ‚ÄĒ most sub-assemblies don't exist yet, so waiting the
                # full timeout on every row is wasted time. quick single-shot
                # pick first; only fall through to _CUSTOMNEW if not present.
                wait_for_options(page, sub_sel)
                options = snapshot_options(page, sub_sel)
                choice = _pick(options, sub_raw)
                if choice is None:
                    # not in dropdown ‚Üí use "- new -" (_CUSTOMNEW) to create it
                    try:
                        page.locator(sub_sel).select_option(value="_CUSTOMNEW")
                        page.locator(sub_sel).dispatch_event("change")
                        name_sel = None
                        for cand in (
                            "#DTE_Field_sub_assembly_name",
                            "#DTE_Field_subassembly_name",
                        ):
                            try:
                                page.wait_for_selector(
                                    cand, state="visible", timeout=3000
                                )
                                name_sel = cand
                                break
                            except Exception:
                                continue
                        if name_sel is None:
                            raise RuntimeError(
                                "sub-assembly name input never became visible"
                            )
                        page.locator(name_sel).fill(sub_raw)
                        log(
                            f"Row {item['row']}: Sub-assembly '{sub_raw}' not found ‚ÄĒ creating new",
                            "OK",
                        )
                    except Exception as e:
                        log(
                            f"Row {item['row']}: Failed to create sub-assembly '{sub_raw}' ({e}) ‚ÄĒ leaving blank",
                            "WARN",
                        )
                else:
                    page.locator(sub_sel).select_option(label=choice)
                    page.locator(sub_sel).dispatch_event("change")
            elif sub_raw and not sub_sel:
                log(
                    f"Row {item['row']}: Sub-assembly dropdown not found in DOM ‚ÄĒ leaving blank",
                    "WARN",
                )

            page.locator("#DTE_Field_part").fill(part_name)

            if item["makebuy"] == "m":
                page.locator("#DTE_Field_makebuy_0").check()
            else:
                page.locator("#DTE_Field_makebuy_1").check()

            if item["comments"]:
                page.locator("#DTE_Field_comments").fill(item["comments"])
            if item["quantity"]:
                page.locator("#DTE_Field_quantity").fill(item["quantity"])

            if DRY_RUN:
                log(f"Row {item['row']}: [DRY RUN] Would create '{part_name}'", "DRY")
                page.wait_for_timeout(DRY_RUN_HOLD_MS)
                close_modal()
            else:
                page.get_by_text("Create", exact=True).click()
                page.wait_for_selector(
                    ".DTE_Action_Create", state="hidden", timeout=10000
                )
                log(f"Row {item['row']}: ‚úď '{part_name}'", "OK")

        for item in filtered:
            sys_code = item["system"]
            asm_raw = item["assembly"]
            sub_raw = item["subassembly"]
            part_name = item["part"]
            row_num = item["row"]

            dup_key = f"{sys_code}_{asm_raw}_{sub_raw}_{part_name}".lower()
            if dup_key in existing:
                log(f"Row {row_num}: Duplicate ‚ÄĒ '{part_name}' already exists", "SKIP")
                skipped_dup += 1
                continue

            last_err: Exception | None = None
            for attempt in range(2):
                try:
                    try_create(item)
                    existing.add(dup_key)
                    success += 1
                    last_err = None
                    break
                except Exception as e:
                    last_err = e
                    close_modal()
                    if attempt == 0:
                        log(f"Row {row_num}: retrying after error ‚ÄĒ {e}", "WARN")
                        page.wait_for_timeout(1000)

            if last_err is not None:
                log(f"Row {row_num}: ‚úó '{part_name}' ‚ÄĒ {last_err}", "ERROR")
                log_error_csv(item, str(last_err))
                failed += 1

        elapsed = round(time.time() - start_time, 1)
        log("‚ĒÄ" * 60)
        log(
            f"Done in {elapsed}s ‚ÄĒ {success} uploaded / {skipped_dup} duplicates / {failed} failed"
        )
        log("‚ĒÄ" * 60)

        input("\nPress ENTER to close the browser...")
        browser.close()
